import threading
import cv2
from datetime import datetime
import numpy as np
from pyzbar.pyzbar import decode
import openpyxl as xl

hora_categorias = {
    "Mañana": {"inicio": 6, "fin": 12},
    "Tarde": {"inicio": 12, "fin": 18},
    "Noche": {"inicio": 18, "fin": 24}
}

cap = cv2.VideoCapture(0)
mañana = set()
tarde = set()
noche = set()

# Define a semaphore with an initial value
semaphore = threading.Semaphore(1)  # Set the initial value to 1


def infhora():
    inf = datetime.now()
    fecha = inf.strftime('%Y:%m:%d')
    hora = inf.strftime('%H:%M:%S')
    return hora, fecha


def crear_archivo_excel(hora_categorias):
    wb = xl.Workbook()
    for categoria in hora_categorias:
        wb.create_sheet(categoria)
    wb.save('RegistroQR.xlsx')


def inicializar_semaforo():
    crear_archivo_excel(hora_categorias)
    print("Semaphore initialized.")


def obtener_categoria_hora(hora):
    for categoria, rango in hora_categorias.items():
        if rango["inicio"] <= hora < rango["fin"]:
            return categoria
    return "Noche"


def actualizar_semaforo_excel(codigo, categoria):
    with semaphore:
        print(f"Thread {threading.current_thread().name} acquired the semaphore.")
        wb = xl.load_workbook('RegistroQR.xlsx')

        # Check if the worksheet exists, create it if not
        if categoria not in wb.sheetnames:
            wb.create_sheet(categoria)

        hoja = wb[categoria]

        # Verificar si el código ya ha sido registrado
        if codigo not in mañana and codigo not in tarde and codigo not in noche:
            # Agregar a la lista correspondiente
            if categoria == "Mañana":
                mañana.add(codigo)
            elif categoria == "Tarde":
                tarde.add(codigo)
            elif categoria == "Noche":
                noche.add(codigo)

            # Agregar a la hoja de Excel
            hoja.append([codigo, infhora()[0]])

            print(f"Added information to Excel: {codigo} in {categoria}")

        print(f"Thread {threading.current_thread().name} released the semaphore.")
        wb.save('RegistroQR.xlsx')


# Empezamos
inicializar_semaforo()

while True:
    ret, frame = cap.read()

    cv2.putText(frame, 'Locate the QR Code', (160, 80), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
    cv2.rectangle(frame, (170, 100), (400, 470), (0, 255, 0), 2)

    hora, fecha = infhora()
    diasem = datetime.today().weekday()

    a, me, d = fecha[0:4], fecha[5:7], fecha[8:10]
    h, m, s = int(hora[0:2]), int(hora[3:5]), int(hora[6:8])

    for codes in decode(frame):
        info = codes.data.decode('utf-8')
        tipo = info[0:2]
        tipo = int(tipo)
        letr = chr(tipo)
        num = info[2:]
        pts = np.array([codes.polygon], np.int32)
        xi, yi = codes.rect.left, codes.rect.top
        pts = pts.reshape((-1, 1, 2))
        codigo = letr + num

        categoria_hora_actual = obtener_categoria_hora(h)

        cv2.polylines(frame, [pts], True, (255, 255, 0), 5)
        actualizar_semaforo_excel(codigo, categoria_hora_actual)

        cv2.putText(frame, letr + '0' + str(num), (xi, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)

    cv2.imshow(" LECTOR DE QR", frame)

    t = cv2.waitKey(5)
    if t == 27:
        break

cv2.destroyAllWindows()
cap.release()

